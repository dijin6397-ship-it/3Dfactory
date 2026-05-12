import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageDraw, ImageFont
import os

SRC = r'F:\自开发程序\工厂建模\图片\规划图.xlsx'
DST = r'F:\自开发程序\工厂建模\图片\规划图.jpg'

PIXELS_PER_CHAR_WIDTH = 8
DEFAULT_ROW_HEIGHT = 20
PADDING = 6
DPI_SCALE = 2
BORDER_COLOR = (0, 0, 0)
BG_COLOR = (255, 255, 255)

def get_merged_map(ws):
    merged_map = {}
    for mc in ws.merged_cells.ranges:
        min_row, min_col = mc.min_row, mc.min_col
        max_row, max_col = mc.max_row, mc.max_col
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if (r, c) != (min_row, min_col):
                    merged_map[(r, c)] = (min_row, min_col)
        merged_map.setdefault((min_row, min_col), None)
    return merged_map, {mc.min_row: mc for mc in ws.merged_cells.ranges}

def color_to_rgb(color_obj):
    if color_obj is None:
        return None
    if color_obj.type == 'rgb' and color_obj.rgb:
        rgb_hex = str(color_obj.rgb)
        if len(rgb_hex) == 8:
            return tuple(int(rgb_hex[i:i+2], 16) for i in (2, 4, 6))
        elif len(rgb_hex) == 6:
            return tuple(int(rgb_hex[i:i+2], 16) for i in (0, 2, 4))
    if color_obj.type == 'theme':
        theme_colors = {
            0: (255, 255, 255), 1: (0, 0, 0), 2: (44, 62, 80),
            3: (255, 255, 255), 4: (0, 176, 80), 5: (255, 0, 0),
            6: (0, 112, 192), 7: (255, 255, 0), 8: (0, 176, 240),
        }
        base = theme_colors.get(color_obj.theme, (255, 255, 255))
        return base
    if color_obj.type == 'indexed':
        indexed_colors = {
            0: (0, 0, 0), 1: (255, 255, 255), 2: (255, 0, 0),
            3: (0, 255, 0), 4: (0, 0, 255), 5: (255, 255, 0),
            6: (255, 0, 255), 7: (0, 255, 255), 8: (0, 0, 0),
            64: (0, 0, 0),
        }
        return indexed_colors.get(color_obj.indexed, None)
    return None

def get_font(size, bold=False):
    font_size = max(10, int(size * DPI_SCALE * 0.9))
    candidates = [
        r'C:\Windows\Fonts\msyh.ttc',
        r'C:\Windows\Fonts\simhei.ttf',
        r'C:\Windows\Fonts\simsun.ttc',
    ]
    for fp in candidates:
        if os.path.exists(fp):
            try:
                return ImageFont.truetype(fp, font_size, index=0)
            except Exception:
                continue
    return ImageFont.load_default()

def xlsx_to_jpg(src_path, dst_path):
    wb = openpyxl.load_workbook(src_path)
    ws = wb[wb.sheetnames[0]]

    merged_map, merged_ranges = get_merged_map(ws)

    col_widths = {}
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        w = ws.column_dimensions[letter].width
        col_widths[c] = (w if w else 8.0) * PIXELS_PER_CHAR_WIDTH * DPI_SCALE

    row_heights = {}
    for r in range(1, ws.max_row + 1):
        h = ws.row_dimensions[r].height
        row_heights[r] = ((h if h else DEFAULT_ROW_HEIGHT) * DPI_SCALE * 1.3)

    col_x = [0]
    for c in range(1, ws.max_column + 1):
        col_x.append(col_x[-1] + col_widths[c])

    row_y = [0]
    for r in range(1, ws.max_row + 1):
        row_y.append(row_y[-1] + row_heights[r])

    total_w = int(col_x[-1]) + 2
    total_h = int(row_y[-1]) + 2

    img = Image.new('RGB', (total_w, total_h), BG_COLOR)
    draw = ImageDraw.Draw(img)

    for mc in ws.merged_cells.ranges:
        x1 = int(col_x[mc.min_col - 1]) + 1
        y1 = int(row_y[mc.min_row - 1]) + 1
        x2 = int(col_x[mc.max_col]) - 1
        y2 = int(row_y[mc.max_row]) - 1
        draw.rectangle([x1, y1, x2, y2], outline=BORDER_COLOR, width=DPI_SCALE)

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if (r, c) in merged_map and merged_map[(r, c)] is not None:
                continue

            x1 = int(col_x[c - 1]) + 1
            y1 = int(row_y[r - 1]) + 1
            x2 = int(col_x[c]) - 1
            y2 = int(row_y[r]) - 1

            if (r, c) not in merged_map:
                draw.rectangle([x1, y1, x2, y2], outline=BORDER_COLOR, width=DPI_SCALE)

            cell = ws.cell(r, c)

            bg = color_to_rgb(cell.fill.fgColor) if cell.fill and cell.fill.fill_type else None
            if bg and bg != (0, 0, 0) and bg != (255, 255, 255):
                draw.rectangle([x1 + 1, y1 + 1, x2 - 1, y2 - 1], fill=bg)

            if cell.value is not None:
                font_size = cell.font.size if cell.font.size else 11
                bold = cell.font.bold if cell.font.bold else False
                pil_font = get_font(font_size, bold)

                text = str(cell.value)
                text_color = color_to_rgb(cell.font.color)
                if text_color is None:
                    text_color = (0, 0, 0)

                if (r, c) in merged_map and merged_map[(r, c)] is None:
                    merge_key = (r, c)
                    for mc in ws.merged_cells.ranges:
                        if mc.min_row == r and mc.min_col == c:
                            x1 = int(col_x[mc.min_col - 1]) + 1
                            y1 = int(row_y[mc.min_row - 1]) + 1
                            x2 = int(col_x[mc.max_col]) - 1
                            y2 = int(row_y[mc.max_row]) - 1
                            break

                cell_w = x2 - x1
                cell_h = y2 - y1

                bbox = draw.textbbox((0, 0), text, font=pil_font)
                text_w = bbox[2] - bbox[0]
                text_h = bbox[3] - bbox[1]

                h_align = cell.alignment.horizontal if cell.alignment.horizontal else 'left'
                v_align = cell.alignment.vertical if cell.alignment.vertical else 'center'

                if h_align == 'center':
                    tx = x1 + (cell_w - text_w) // 2
                elif h_align == 'right':
                    tx = x2 - text_w - PADDING * DPI_SCALE
                else:
                    tx = x1 + PADDING * DPI_SCALE

                if v_align == 'center':
                    ty = y1 + (cell_h - text_h) // 2
                elif v_align == 'bottom':
                    ty = y2 - text_h - PADDING * DPI_SCALE
                else:
                    ty = y1 + PADDING * DPI_SCALE

                if cell.alignment.wrap_text and text_w > cell_w - PADDING * 2 * DPI_SCALE:
                    lines = []
                    current_line = ""
                    for char in text:
                        test_line = current_line + char
                        bbox_t = draw.textbbox((0, 0), test_line, font=pil_font)
                        if bbox_t[2] - bbox_t[0] > cell_w - PADDING * 2 * DPI_SCALE:
                            if current_line:
                                lines.append(current_line)
                            current_line = char
                        else:
                            current_line = test_line
                    if current_line:
                        lines.append(current_line)
                    
                    total_text_h = len(lines) * (text_h + 2 * DPI_SCALE)
                    if v_align == 'center':
                        ty = y1 + (cell_h - total_text_h) // 2
                    
                    for line in lines:
                        draw.text((tx, ty), line, fill=text_color, font=pil_font)
                        ty += text_h + 2 * DPI_SCALE
                else:
                    draw.text((tx, ty), text, fill=text_color, font=pil_font)

    img = img.convert('RGB')
    img.save(dst_path, 'JPEG', quality=95, dpi=(300, 300))
    print(f"Saved: {dst_path}")
    print(f"Size: {total_w}x{total_h} pixels")
    wb.close()

if __name__ == '__main__':
    xlsx_to_jpg(SRC, DST)
